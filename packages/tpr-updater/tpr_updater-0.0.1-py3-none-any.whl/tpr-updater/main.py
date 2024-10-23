import pandas as pd
from convert import convert_xls_to_xlsx
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkinter import filedialog, messagebox

def process_tpr_file(file_path):

    print("Processing file: ", file_path)

    book = load_workbook(file_path)
    sheet_name = 'TPR Items'  
    sheet = book[sheet_name]

    # Load the data into a pandas DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=[5, 6], engine='openpyxl')

    # Combine the multi-row headers into a single row
    df.columns = [' '.join(col).strip() for col in df.columns.values]

    # Perform your data modifications in pandas
    for index, row in df.iterrows():
        # Update TPR SRP if SRP is not 0 and less than TPR SRP
        if row['SRP SRP'] != 0 and row['SRP SRP'] < row['TPR SRP']:
            df.loc[index, 'TPR SRP'] = row['SRP SRP']
            # Mark this cell for formatting later
            sheet.cell(row=index+8, column=df.columns.get_loc('TPR SRP')+1).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            print(f"Updated row {index+8}, column 'TPR SRP' from {row['TPR SRP']} to {row['SRP SRP']}")

    # Write the modified DataFrame back to the same sheet
    for r_idx, row in enumerate(df.values, 8):  # Start from row 8 (7 + 1 for 0-indexing)
        for c_idx, value in enumerate(row, 1):  # Start from column 1
            sheet.cell(row=r_idx, column=c_idx, value=value)

    output_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if output_file_path:
        book.save(output_file_path)
        messagebox.showinfo("Success", f"Updated data saved to {output_file_path}")


if __name__ == "__main__":

    print("Please select the TPR file to process")

    # time.sleep(5)

    # Load the Excel file using openpyxl (preserves formatting)
    file_path = filedialog.askopenfilename(filetypes=[("All files", "*.*")])

    # if the file is an xls file, convert it to xlsx
    if file_path.endswith('.xls'):
        file_path = convert_xls_to_xlsx(file_path)
    elif not file_path:
        messagebox.showerror("Error", "No file selected. Exiting program.")
        exit()
    else:
        messagebox.showerror("Error", "File is not an xls or xlsx file. Exiting program.")
        exit()

    process_tpr_file(file_path)
