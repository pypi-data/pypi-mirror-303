import openpyxl as xl


class WorksheetDumpControl():


    @staticmethod
    def dump(worksheet, file):
        ws = worksheet

        with open(file, mode='w', encoding='utf8') as f:
            for row_th in range(1, ws.max_row + 1):
                for column_th in range(1, ws.max_column + 1):

                    column_letter = xl.utils.get_column_letter(column_th)
                    cell = ws[f'{column_letter}{row_th}']

                    items = []

                    # とりあえず罫線を出力
                    if cell.border is not None:
                        if hasattr(cell.border, 'left') and cell.border.left is not None:
                            sub_items = WorksheetDumpControl.parse_side(side=cell.border.left)
                            if 0 < len(sub_items):
                                items.append(f"left={' '.join(sub_items)}")

                        if hasattr(cell.border, 'right') and cell.border.right is not None:
                            sub_items = WorksheetDumpControl.parse_side(side=cell.border.right)
                            if 0 < len(sub_items):
                                items.append(f"right={' '.join(sub_items)}")

                        if hasattr(cell.border, 'top') and cell.border.top is not None:
                            sub_items = WorksheetDumpControl.parse_side(side=cell.border.top)
                            if 0 < len(sub_items):
                                items.append(f"top={' '.join(sub_items)}")

                        if hasattr(cell.border, 'bottom') and cell.border.bottom is not None:
                            sub_items = WorksheetDumpControl.parse_side(side=cell.border.bottom)
                            if 0 < len(sub_items):
                                items.append(f"bottom={' '.join(sub_items)}")

                        if hasattr(cell.border, 'diagonal') and cell.border.diagonal is not None:
                            sub_items = WorksheetDumpControl.parse_side(side=cell.border.diagonal)
                            if 0 < len(sub_items):
                                items.append(f"diagonal={' '.join(sub_items)}")

                        if hasattr(cell.border, 'vertical') and cell.border.vertical is not None:
                            sub_items = WorksheetDumpControl.parse_side(side=cell.border.vertical)
                            if 0 < len(sub_items):
                                items.append(f"vertical={' '.join(sub_items)}")

                        if hasattr(cell.border, 'horizontal') and cell.border.horizontal is not None:
                            sub_items = WorksheetDumpControl.parse_side(side=cell.border.horizontal)
                            if 0 < len(sub_items):
                                items.append(f"horizontal={' '.join(sub_items)}")

                    f.write(f"""\
({row_th}, {column_th}) {' '.join(items)}
""")


    @staticmethod
    def parse_side(side):
        items = []
        if side.style is not None:
            items.append(f"{side.style}")
        # TODO Color 型は RPG形式以外もあるが、必要になったら対応する
        if side.color is not None:
            items.append(f"{side.color.rgb}")
        return items
