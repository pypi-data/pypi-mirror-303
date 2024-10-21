import os
import datetime
import openpyxl as xl

from .database import Table
from .workbooks import TreeDrawer, TreeEraser


class Settings():
    """各種設定"""


    def __init__(self, dictionary=None):
        """初期化
        
        Parameters
        ----------
        dictionary : dict
            設定

            列の幅設定。width はだいたい 'ＭＳ Ｐゴシック' サイズ11 の半角英文字の個数
            * `column_width_of_no` - A列の幅。no列
            * `column_width_of_row_header_separator` - B列の幅。空列
            * `column_width_of_node` - 例：C, F, I ...列の幅。ノードの箱の幅
            * `column_width_of_parent_side_edge` - 例：D, G, J ...列の幅。エッジの水平線のうち、親ノードの方
            * `column_width_of_child_side_edge` - 例：E, H, K ...列の幅。エッジの水平線のうち、子ノードの方

            行の高さ設定。height の単位はポイント。既定値 8。昔のアメリカ人が椅子に座ってディスプレイを見たとき 1/72 インチに見える大きさが 1ポイント らしいが、そんなんワカラン。目視確認してほしい
            * `row_height_of_header` - 第１行。ヘッダー
            * `row_height_of_column_header_separator` - 第２行。空行
            * `row_height_of_upper_side_of_node` - ノードの上側のセルの高さ
            * `row_height_of_lower_side_of_node` - ノードの下側のセルの高さ
            * `row_height_of_node_spacing` - ノード間の高さ

            * 色の参考： 📖 [Excels 56 ColorIndex Colors](https://www.excelsupersite.com/what-are-the-56-colorindex-colors-in-excel/)

            背景色関連
            * `bgcolor_of_header_1` - ヘッダーの背景色その１
            * `bgcolor_of_header_2` - ヘッダーの背景色その２
            * `bgcolor_of_node` - 背景色

            文字色関連
            * `fgcolor_of_header_1` - ヘッダーの文字色その１
            * `fgcolor_of_header_2` - ヘッダーの文字色その２

            文字寄せ関連
            * `horizontal_alignment_of_node` - 文字の水平方向の寄せ。規定値 None。'left', 'fill', 'centerContinuous', 'center', 'right', 'general', 'justify', 'distributed' のいずれか。指定しないなら None
            * `vertical_alignment_of_node` - 文字の垂直方向の寄せ。規定値 None。'bottom', 'center', 'top', 'justify', 'distributed' のいずれか。指定しないなら None
        """

        # 既定のディクショナリー
        self._dictionary = {
            # 列の幅
            'column_width_of_no':                         4,
            'column_width_of_row_header_separator':       3,
            'column_width_of_node':                       20,
            'column_width_of_parent_side_edge':           2,
            'column_width_of_child_side_edge':            4,

            # 行の高さ
            'row_height_of_header':                    13,
            'row_height_of_column_header_separator':   13,
            'row_height_of_upper_side_of_node':     13,
            'row_height_of_lower_side_of_node':     13,
            'row_height_of_node_spacing':           6,

            # 背景色関連
            'bgcolor_of_header_1':                 'CCCCCC',
            'bgcolor_of_header_2':                 '333333',
            'bgcolor_of_node':                     'FFFFCC',

            # 文字色関連
            'fgcolor_of_header_1':                 '111111',
            'fgcolor_of_header_2':                 'EEEEEE',

            # 文字寄せ関連
            'horizontal_alignment_of_node':        None,
            'vertical_alignment_of_node':          None,
        }

        # 上書き
        if dictionary is not None:
            for key, value in dictionary.items():
                self._dictionary[key] = value


    @property
    def dictionary(self):
        return self._dictionary


class SettingsOfNode():
    """TODO ノード個別の設定"""


    def __init__(self, dictionary=None):
        """初期化
        
        Parameters
        ----------
        dictionary : dict
            設定

            * 色の参考： 📖 [Excels 56 ColorIndex Colors](https://www.excelsupersite.com/what-are-the-56-colorindex-colors-in-excel/)

            色関連
            * `bgcolor` - 背景色
            * `fgcolor` - 文字色

            文字寄せ関連
            * `horizontal_alignment` - 文字の水平方向の寄せ。規定値 None。'left', 'fill', 'centerContinuous', 'center', 'right', 'general', 'justify', 'distributed' のいずれか。指定しないなら None
            * `vertical_alignment_` - 文字の垂直方向の寄せ。規定値 None。'bottom', 'center', 'top', 'justify', 'distributed' のいずれか。指定しないなら None
        """

        # 既定のディクショナリー
        self._dictionary = {

            # 色関連
            'bgcolor':                     'FFFFCC',
            'fgcolor':                     None,

            # 文字寄せ関連
            'horizontal_alignment':        None,
            'vertical_alignment':          None,
        }

        # 上書き
        if dictionary is not None:
            for key, value in dictionary.items():
                self._dictionary[key] = value


    @property
    def dictionary(self):
        return self._dictionary


class WorkbookControl():
    """ワークブック制御"""


    def __init__(self, target, mode, settings={}, debug_write=False):
        """初期化

        Parameters
        ----------
        target : str
            ワークブック（.xlsx）へのファイルパス
        mode : str
            既存のワークブックが有ったときの挙動。 'w' は新規作成して置換え、 'a' は追記
        settings : dict
            各種設定
        """
        self._wb_file_path = target
        self._mode = mode
        self._settings_obj = Settings(dictionary=settings)
        self._debug_write = debug_write
        self._wb = None
        self._ws = None


    @property
    def workbook_file_path(self):
        return self._wb_file_path


    def render_worksheet(self, target, based_on, debug_write=False):
        """ワークシート描画

        Parameters
        ----------
        target : str
            シート名
        based_on : str
            CSVファイルパス
        debug_write : bool
            デバッグライト
        """

        if self._wb is None:
            self.ready_workbook()

        self.ready_worksheet(target=target)

        # CSV読込
        table = Table.from_csv(file_path=based_on)

        # ツリードロワーを用意、描画（都合上、要らない罫線が付いています）
        tree_drawer = TreeDrawer(table=table, ws=self._ws, settings_obj=self._settings_obj, debug_write=debug_write)
        tree_drawer.render()


        # 要らない罫線を消す
        # DEBUG_TIPS: このコードを不活性にして、必要な線は全部描かれていることを確認してください
        if True:
            tree_eraser = TreeEraser(table=table, ws=self._ws, debug_write=debug_write)
            tree_eraser.render()
        else:
            if self._debug_write:
                print(f"[{datetime.datetime.now()}] eraser disabled")


    def remove_worksheet(self, target):
        """存在すれば、指定のワークシートの削除。存在しなければ無視

        Parameters
        ----------
        target : str
            シート名
        """

        if self.exists_sheet(target=target):
            if self._debug_write:
                print(f"[{datetime.datetime.now()}] remove `{target}` sheet...")

            self._wb.remove(self._wb[target])


    def save_workbook(self):
        """ワークブックの保存"""

        if self._debug_write:
            print(f"[{datetime.datetime.now()}] save `{self._wb_file_path}` file...")

        # ワークブックの保存            
        self._wb.save(self._wb_file_path)


    def ready_workbook(self):
        """ワークブックを準備する"""

        # 既存のファイルが有ったときの挙動
        if os.path.isfile(self._wb_file_path):
            # 既存のファイルへ追記
            if self._mode == 'a':
                if self._debug_write:
                    print(f"[{datetime.datetime.now()}] `{self._wb_file_path}` file exists, read.")

                self._wb = xl.load_workbook(filename=self._wb_file_path)

                return
            
            elif self._mode == 'w':
                pass
            
            else:
                raise ValueError(f"unsupported {self._mode=}")

                    
        # ワークブックを新規生成
        if self._debug_write:
            print(f"[{datetime.datetime.now()}] `{self._wb_file_path}` file not exists, create.")

        self._wb = xl.Workbook()


    def exists_sheet(self, target):
        """シートの存在確認
        
        Parameters
        ----------
        target : str
            シート名
        """
        return target in self._wb.sheetnames


    def ready_worksheet(self, target):
        """ワークシートを準備する
        
        Parameters
        ----------
        target : str
            シート名
        """

        # シートを作成
        if not self.exists_sheet(target):
            if self._debug_write:
                print(f"[{datetime.datetime.now()}] create `{target}` sheet...")

            self._wb.create_sheet(target)


        self._ws = self._wb[target]
