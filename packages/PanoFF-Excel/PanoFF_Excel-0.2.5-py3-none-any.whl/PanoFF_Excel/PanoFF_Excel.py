import random
import openpyxl
from openpyxl.styles import PatternFill
import xlrd
from openpyxl import Workbook
import os


def xls_to_xlsx(xls_file, xlsx_file):
    workbook_xls = xlrd.open_workbook(xls_file)
    sheet_xls = workbook_xls.sheet_by_index(0)

    workbook_xlsx = Workbook()
    sheet_xlsx = workbook_xlsx.active

    # Copy data from .xls to .xlsx
    for row in range(sheet_xls.nrows):
        for col in range(sheet_xls.ncols):
            sheet_xlsx.cell(row=row + 1, column=col + 1, value=sheet_xls.cell_value(row, col))  # Fixed column argument

    # Save the new .xlsx file
    workbook_xlsx.save(xlsx_file)

    os.remove(xls_file)


# Usage
def read_xlsx(filepath):
    if filepath.lower().endswith('.xls'):
        xlsx_filepath = f"{os.path.splitext(filepath)[0]}.xlsx"
        xls_to_xlsx(filepath, xlsx_filepath)
        filepath = xlsx_filepath

    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active

    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def random_color():
    """Генерирует случайный цвет в формате RGB"""
    return f"{random.randint(0, 255):02X}{random.randint(0, 255):02X}{random.randint(0, 255):02X}"


def adjust_column_widths(sheet):
    # Проходим по всем столбцам
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Получаем букву столбца

        # Находим максимальную длину текста в колонке
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        # Устанавливаем ширину столбца с небольшим запасом (например, +2)
        adjusted_width = (max_length + 2) * 1.1  # Коэффициент 1.1 для большей точности
        sheet.column_dimensions[column].width = adjusted_width


def write_xlsx(filepath, data, fills=None):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for i, row in enumerate(data):
        for j, value in enumerate(row):
            cell = sheet.cell(row=i + 1, column=j + 1, value=value)

            try:
                # Устанавливаем заливку, если в fills есть информация для этой ячейки
                if (i, j) in fills:
                    color = fills[(i, j)]
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            except Exception as e:
                print(e)

    # Применяем настройку ширины колонок после заполнения данных
    adjust_column_widths(sheet)

    workbook.save(filepath)


def numeric(data):
    # Обновляем заголовок
    header_row = data[0][:3:] + list(range(1, 51))
    data[0] = header_row

    # Обработка данных
    for i, row in enumerate(data):
        # Проверка для первой строки (индекс 0)
        if i == 0:
            # Проверка на None в индексе 3
            if row[3] is None:
                row[3] = i - 1  # Заменяем None на индекс строки
        elif i > 0:
            for j in range(0, 50):
                data[i].append(None)
    return data


def get_full_table(data):
    data_len = len(data)
    fills = None
    for i in range(1, data_len):
        current_data_line = data[i]

        row_color = random_color()

        # Если нет зависимостей (ID процессов A равен 0)
        if current_data_line[2] == 0:
            start_index = 3
            for j in range(start_index, start_index + current_data_line[1]):
                if j < len(current_data_line) and current_data_line[j] is None:
                    current_data_line[j] = 1
                    fills[(i, j)] = row_color
            data[i] = current_data_line  # Обновляем текущую строку в data

        else:
            # Обработка зависимостей
            try:
                current_data_line[2] = str(current_data_line[2]).replace(' ', '').strip(';')
                ids = [int(x) for x in current_data_line[2].split(';') if x.isdigit()]

                # Инициализация для поиска максимального индекса
                max_index = -1

                # Находим максимальный индекс среди зависимых процессов
                for pid in ids:
                    for j in range(1, data_len):
                        if data[j][0] == pid:  # Проверяем, совпадает ли ID процесса
                            # Находим последний индекс с 1
                            for k in range(3, len(data[j])):
                                if data[j][k] == 1:
                                    max_index = max(max_index, k)

                # Если мы нашли максимальный индекс, заполняем 1
                if max_index >= 0:
                    start_index = max_index + 1
                    for j in range(start_index, start_index + current_data_line[1]):
                        if j < len(current_data_line) and current_data_line[j] is None:
                            current_data_line[j] = 1
                            fills[(i, j)] = row_color

                data[i] = current_data_line  # Обновляем текущую строку в data

            except Exception as e:
                print(f"Ошибка при обработке: {e}")

    data = convert_strings_to_int(data)
    return data, fills


def create(filepath):
    data_main = read_xlsx(filepath)
    numeric_data = numeric(data_main)
    data, fills = get_full_table(numeric_data)
    write_xlsx(filepath="output.xlsx", data=data, fills=fills)


def convert_strings_to_int(data):
    # Проходим по каждому элементу массива
    for i in range(len(data)):
        # Если элемент - это список, рекурсивно обрабатываем его
        if isinstance(data[i], list):
            convert_strings_to_int(data[i])
        else:
            # Если элемент является строкой и представляет собой число, преобразуем его в int
            if isinstance(data[i], str) and data[i].isdigit():
                data[i] = int(data[i])
            # Если элемент это строка, но с возможным знаком "-", преобразуем корректно
            elif isinstance(data[i], str):
                try:
                    data[i] = int(data[i])
                except ValueError:
                    pass  # Оставляем как есть, если строка не может быть преобразована в число

    return data


def solving_the_problem(data):
    fills = {}
    data_len = len(data)
    for i in range(1, data_len):
        current_data_line = data[i]

        if current_data_line[2] == 0:
            current_data_line[3] = current_data_line[1]

        else:
            try:
                current_data_line[2] = str(current_data_line[2]).replace(' ', '').strip(';')
                ids = [int(x) for x in current_data_line[2].split(';') if x.isdigit()]

                ms = []

                for pid in ids:
                    if len(ids) == 2:
                        for j in range(1, data_len):
                            if data[j][0] == pid:
                                k = data[pid][3]
                                ms.append(k)
                                max_index = max(ms)
                                current_data_line[3] = current_data_line[1] + max_index
                    if len(ids) == 1:
                        schet = int(current_data_line[1]) + data[int(current_data_line[2])][3]
                        current_data_line[3] = schet

            except Exception as e:
                print(f"Ошибка при обработке: {e}")
        data[i] = current_data_line

        data = convert_strings_to_int(data)

    return data, fills


def decision(filepath):
    try:
        data_main = read_xlsx(filepath)
        data, fills = solving_the_problem(data_main)
        write_xlsx(filepath="output.xlsx", data=data, fills=fills)
    except Exception as e:
        print(f"Error in processing {filepath}: {e}")
