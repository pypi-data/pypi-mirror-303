import collections
import datetime
import math
import os
import sys
import typing
import unicodedata
from pathlib import Path

import pandas as pd
import yaml
from openpyxl import Workbook, formatting, load_workbook
from openpyxl.styles import Alignment, Border, Font, GradientFill, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from PIL import Image

from NexR_qc.Logging import *
from NexR_qc.Timer import *


class ImageQC:

    def __init__(self, DataName=None, Mode="일반", PathDict=None, logger_save=False):

        if DataName is not None:
            self.DataName = DataName

        self.Mode = Mode

        # 디렉토리 세팅
        if PathDict:
            self.PATH = PathDict
        else:
            self.PATH = {}
            self.PATH.setdefault("ROOT", Path.cwd())
            if self.DataName is not None:
                self.PATH.setdefault("DATA", self.PATH["ROOT"] / "data" / self.DataName)
            else:
                self.PATH.setdefault("DATA", self.PATH["ROOT"] / "data")
            self.PATH.setdefault("IMAGE", self.PATH["DATA"] / "images")
            if self.Mode == "객체탐지":
                self.PATH.setdefault("LABEL", self.PATH["DATA"] / "labels")
                self.PATH.setdefault("YAML", self.PATH["ROOT"] / f"{self.DataName}.yaml")

        self.PATH.setdefault("LOG", self.PATH["ROOT"] / "log" / self.DataName)
        if self.DataName is not None:
            self.PATH.setdefault("DOCUMENTS", self.PATH["ROOT"] / "documents" / self.DataName)
            self.PATH.setdefault("OUTPUT", self.PATH["ROOT"] / "output" / self.DataName)
        else:
            self.PATH.setdefault("DOCUMENTS", self.PATH["ROOT"] / "documents")
            self.PATH.setdefault("OUTPUT", self.PATH["ROOT"] / "output")

        # 없는 디렉토리 생성
        for k in self.PATH:
            if not k in ["CURRENT", "CATEGORY", "DATA", "IMAGE", "LABEL", "YAML"]:
                Path.mkdir(self.PATH[k], parents=True, exist_ok=True)

        # 로그 파일 세팅
        self.logger = Logger(
            proc_name="QualityCheck",
            log_folder_path=self.PATH["LOG"],
            save=logger_save,
        )

        # 로그 출력 색상값 세팅
        self.colorSetting = {"grey": "\x1b[38;20m", "blue": "\033[34m", "green": "\033[32m", "yellow": "\x1b[33;20m", "red": "\x1b[31;20m", "bold_red": "\x1b[31;1m", "reset": "\x1b[0m"}

        # 실행 시간 타이머 설정
        self.timer = Timer(logger=self.logger)
        self.timer.start()

        if not self.PATH["DATA"].exists():
            self.logger.error(f"데이터 폴더 경로가 존재하지 않습니다. 데이터 폴더 경로를 다시 확인 바랍니다.")
            self.logger.error(f'데이터 폴더 경로: {self.colorSetting["yellow"]} {self.PATH["DATA"]} {self.colorSetting["reset"]}')
            raise FileExistsError

        if self.Mode == "객체탐지":
            if not self.PATH["YAML"].exists():
                self.logger.error(f"yaml 파일이 존재하지 않습니다. yaml 파일 경로를 다시 확인 바랍니다.")
                self.logger.error(f'yaml 파일 경로: {self.colorSetting["yellow"]} {self.PATH["YAML"]} {self.colorSetting["reset"]}')
                raise FileExistsError
            else:
                # 어노테이션 파일 로드
                with open(self.PATH["YAML"], "r") as f:
                    self.yaml_info = yaml.full_load(f)

        # 파일확장자명 정의
        self.ext_fmt = [".png", ".jpg", ".jpeg", ".gif"]

        # 실행 시작 시간
        CreatedTime = datetime.today()
        self.CreatedTime = CreatedTime.strftime("%Y%m%d_%H%M%S")

        # 정보값 저장을 위한 딕셔너리 생성
        self.ResultDict = {"Report": {}, "QCinfo": {}}

    # Step 1: 데이터 현황 메타 정보 테이블 확인
    # Case 1: 이미지 분류 모델 활용 데이터
    def get_file_meta_table(self):
        """
        데이터 파일 메타 정보를 획득한다 (단, 기존 메타 정보가 없을 경우 생성한다)
        """
        self.logger.info(f"{self.colorSetting['blue']}[Step 1] 데이터 메타 정보 테이블 파일 확인{self.colorSetting['reset']}")

        self.logger.info(f'데이터 메타 정보 테이블 확인 폴더 경로: {self.PATH["DOCUMENTS"]}')

        # 기존 데이터 메타 정보 테이블이 존재하는 경우
        if len([path for path in self.PATH["DOCUMENTS"].glob("*") if all([path.suffix.lower() == ".csv", self.Mode in unicodedata.normalize("NFC", path.stem)])]) > 0:
            self.file_meta_path = sorted([path for path in self.PATH["DOCUMENTS"].glob(f"file_meta_info_*") if all([path.suffix.lower() == ".csv", self.Mode in unicodedata.normalize("NFC", path.stem)])], key=lambda x: x.stat().st_mtime)[-1]
            self.logger.info(f"기존 저장된 데이터 메타 정보 테이블을 활용합니다. (파일 경로: {self.file_meta_path})")

            self.file_meta_table = pd.read_csv(self.file_meta_path, converters={"파일명": str}, encoding="utf-8-sig")

        # 기존 데이터 메타 정보 테이블이 없어 생성해야 하는 경우
        else:
            self.logger.info(f"현재 데이터 메타 정보 테이블이 존재하지 않습니다.")
            try:
                # 데이터 메타 정보 테이블 생성 질문
                self.FileMetaInfoCreateQuestion_YN = True if input(f"""{self.colorSetting["yellow"]}데이터 메타 정보 테이블 생성 작업 수행을 원하시면 Y, 원하시지 않다면 N을 입력해주세요 (Y/N):{self.colorSetting["reset"]} """) in ["Y", "y"] else False
                if not self.FileMetaInfoCreateQuestion_YN:
                    raise  # 데이터 메타 정보 테이블이 없는데 생성을 하지 않을 경우 에러 발생

                # 데이터 메타 정보 테이블 생성
                self.logger.info("지금부터 데이터 메타 정보 테이블 생성 작업을 시작합니다.")

                self.file_meta_path = self.PATH["DOCUMENTS"] / f"file_meta_info_{self.Mode}_{self.CreatedTime}.csv"

                self.img_filepath_list = [path for path in self.PATH["IMAGE"].rglob("*") if all([path.suffix.lower() in self.ext_fmt, not path.is_dir()])]  # [일반/분류/객체탐지] 이미지 파일 경로 리스트 획득

                if self.Mode == "분류":
                    self.cate_level_max = max([len(img_path.relative_to(self.PATH["IMAGE"]).parts) - 2 for img_path in self.img_filepath_list])
                    self.cate_cols = [f"분류{str(i+1)}명" for i in range(self.cate_level_max)]

                if self.Mode == "객체탐지":
                    self.label_filepath_list = [path for path in self.PATH["LABEL"].rglob("*") if all([path.suffix.lower() in [".txt"], not path.is_dir()])]  # [객체탐지] 라벨 파일 경로 리스트 획득

                self.filename_list = sorted(list(set([path.stem if not isinstance(path.stem, int) else str(path.stem) for path in self.img_filepath_list])))  # [일반/분류/객체탐지] 이미지 파일명 리스트 정보 획득

                self.logger.info(f"이미지 파일명 개수 : {len(self.filename_list):,}")
                self.logger.info(f"이미지 파일 개수 : {len(self.img_filepath_list):,}")
                if self.Mode == "객체탐지":
                    self.logger.info(f"어노테이션 파일 개수 : {len(self.label_filepath_list):,}")

                if self.Mode == "일반":
                    self.ColList = ["No", "파일명", "데이터명", "이미지파일경로", "너비", "높이"]
                elif self.Mode == "분류":
                    self.ColList = ["No", "파일명", "데이터명", "데이터셋구분"] + self.cate_cols + ["이미지파일경로", "너비", "높이"]
                elif self.Mode == "객체탐지":
                    self.ColList = ["No", "파일명", "데이터명", "데이터셋구분", "이미지파일경로", "라벨파일경로", "너비", "높이", "객체개수"]

                self.file_meta_info = [tuple([idx + 1]) + self.get_file_info(filename) for idx, filename in enumerate(self.filename_list)]
                self.file_meta_table = pd.DataFrame(self.file_meta_info, columns=self.ColList)
                self.file_meta_table.to_csv(self.file_meta_path, index=False, encoding="utf-8-sig")

                self.logger.info(f"데이터 메타 정보 테이블 생성 작업을 완료하였습니다. (파일 생성 경로: {self.file_meta_path})")

            except:
                self.logger.error("데이터 메타 정보 테이블이 없어 QC 수행이 불가합니다.")
                self.logger.error("기존 데이터 메타 정보 테이블이 존재하는 경우 파일 위치를 다시 한번 확인하거나 기존 파일이 없는 경우 생성 작업을 수행바랍니다.")
                raise Exception("에러 발생")

        self.logger.info(f"{self.colorSetting['blue']}[Step 1] 데이터 메타 정보 테이블 파일 확인 완료{self.colorSetting['reset']}")

        return

    def get_file_info(self, filename: str) -> tuple:
        """
        지정된 데이터셋 폴더 내 이미지 데이터 파일 메타 정보를 획득한다

        parameter
        ----------
        path(Union[str, os.PathLike]): 이미지 파일 경로

        return
        ----------
        img_info(tuple): 이미지 파일 메타 정보 (파일명, 데이터셋명, 대분류명, 소분류명, 파일형식, 파일경로, 너비, 높이)

        Usage
        ----------
        dataset_path = './data/img_0000.jpg'
        img_info = get_image_info(dataset_path)
        """

        if isinstance(filename, str):
            pass
        else:
            raise TypeError

        data_name = unicodedata.normalize("NFC", self.DataName)  # [일반/분류/객체탐지] 데이터셋 명

        img_path = [path for path in self.img_filepath_list if filename in str(path)][0]
        if self.Mode in ["분류", "객체탐지"]:
            dataset_gubun = unicodedata.normalize("NFC", list(img_path.relative_to(self.PATH["IMAGE"]).parts)[0])
        if self.Mode in ["분류"]:
            # for i, cate_name in enumerate(img_path.relative_to(self.PATH["IMAGE"]).parts):
            # print(i, cate_name)
            cate_list = [unicodedata.normalize("NFC", cate_name) for i, cate_name in enumerate(img_path.relative_to(self.PATH["IMAGE"]).parts) if all([i > 0, cate_name != img_path.name])]
            cate_tuple = tuple(cate_list + [None] * (self.cate_level_max - len(cate_list)))
        img_path = unicodedata.normalize("NFC", str(img_path.relative_to(self.PATH["ROOT"])))

        if self.Mode in ["객체탐지"]:
            try:
                label_path = [path for path in self.label_filepath_list if filename in str(path)][0]
                label_path = label_path.relative_to(self.PATH["ROOT"])
                with open(label_path, "r") as f:
                    label_info = list(map(lambda x: x.strip().split(), f.readlines()))
                bbox_count = len(label_info)
                label_path = unicodedata.normalize("NFC", str(label_path))
            except:
                self.logger.info(f"{filename} : 이미지 파일 ✅ / 라벨 파일 ⛔️")
                label_path = None

        img = Image.open(img_path)
        width, height = img.size
        filename = unicodedata.normalize("NFC", filename)

        if self.Mode == "일반":
            file_info = (filename, data_name, img_path, width, height)
        elif self.Mode == "분류":
            file_info = (filename, data_name, dataset_gubun) + cate_tuple + (img_path, width, height)
        elif self.Mode == "객체탐지":
            file_info = (filename, data_name, dataset_gubun, img_path, label_path, width, height, bbox_count)

        return file_info

    # 공통 검사항목 (검사항목 001 ~ 003) 검사 실시
    def check_common(self, filename: str) -> list:
        """
        이미지 파일에 대해 아래의 검사항목을 검사합니다.
        - 검사항목 001: 메타 정보 테이블 내 해당 이미지 파일 정보 존재 여부
        - 검사항목 002: 메타 정보 테이블 내 이미지 해상도와 실제 이미지 해상도 일치 여부
        - 검사항목 003: 메타 정보 테이블 내 이미지 해상도와 실제 이미지 해상도 일치 여부

        parameter
        ----------
        filename(str): 파일 확장자를 제외한 이미지 파일명

        return
        ----------
        result(list): 검사항목 001 ~ 003에 대한 검사 결과

        Usage
        ----------
        filename = 'Image_0001'
        result = check_img_exist_meta_table(filename = filename)
        """
        result_001 = filename in self.file_meta_table["파일명"].values
        if result_001:

            meta_info = self.file_meta_table.loc[self.file_meta_table["파일명"] == filename, ["이미지파일경로", "너비", "높이"]].values[0]
            meta_img_path, *meta_img_res = meta_info
            meta_img_path = self.PATH["ROOT"] / Path(meta_img_path)

            current_img_path = self.PATH["ROOT"] / Path(self.ResultDict["QCinfo"].loc[self.ResultDict["QCinfo"]["파일명"] == filename, "이미지파일경로"].values[0])

            try:
                result_002 = all([meta_img_path.is_file(), meta_img_path == current_img_path])
            except:
                result_002 = False

        else:
            result_002 = False

        if result_001:

            img = Image.open(current_img_path)
            current_width, current_height = img.size
            result_003 = meta_img_res == [current_width, current_height]
        else:
            result_003 = False

        result = [int(result_001), int(result_002), int(result_003)]

        return result

    # 분류 검사항목 (검사항목 004 ~ 006) 검사 실시
    def check_classification(self, filename: str) -> list:
        """
        이미지 파일에 대해 아래의 검사항목을 검사합니다.
        - 검사항목 004: 메타 파일 내 데이터셋 구분값과 실제 이미지 파일 분류 현황 일치 여부
        - 검사항목 005: 메타 파일 내 분류값과 실제 이미지 파일 분류 현황 일치 여부

        parameter
        ----------
        filename(str): 파일 확장자를 제외한 이미지 파일명

        return
        ----------
        result(list): 검사항목 004 ~ 006에 대한 검사 결과

        Usage
        ----------
        filename = 'Image_0001'
        result = check_img_exist_meta_table(filename = filename)
        """

        self.cate_cols = [col for col in self.file_meta_table.columns if all(["분류" in col, "명" in col])]
        self.cate_level_max = len(self.cate_cols)
        result_001 = self.ResultDict["QCinfo"].loc[self.ResultDict["QCinfo"]["파일명"] == filename, ["검사_001"]].values[0]

        if bool(result_001):

            meta_info = self.file_meta_table.loc[self.file_meta_table["파일명"] == filename, ["데이터셋구분"] + self.cate_cols].values[0]
            meta_dataset_gubun, *meta_cate_name = meta_info
            meta_cate_name = [cate_name if isinstance(cate_name, str) else None for cate_name in meta_cate_name]

            current_img_path = self.PATH["ROOT"] / Path(self.ResultDict["QCinfo"].loc[self.ResultDict["QCinfo"]["파일명"] == filename, "이미지파일경로"].values[0])

            current_dataset_gubun = unicodedata.normalize("NFC", list(current_img_path.relative_to(self.PATH["IMAGE"]).parts)[0])
            current_cate_list = [cate_name for i, cate_name in enumerate(current_img_path.relative_to(self.PATH["IMAGE"]).parts) if all([i > 0, cate_name != current_img_path.name])]
            current_cate_list = current_cate_list + [None] * (self.cate_level_max - len(current_cate_list))

            result_004 = meta_dataset_gubun == current_dataset_gubun
            result_005 = meta_cate_name == current_cate_list

        else:
            result_004, result_005 = False, False

        result = [int(result_004), int(result_005)]

        return result

    # 객체탐지 검사항목 (검사항목 004 ~ 011) 검사 실시
    def check_object_detection(self, filename: str) -> list:
        """
        이미지 파일에 대해 아래의 검사항목을 검사합니다.
        - 검사항목 004: 메타 파일 내 데이터셋 구분값과 실제 이미지 파일 분류 현황 일치 여부
        - 검사항목 005: 이미지 파일에 대한 라벨 파일 존재 여부
        - 검사항목 006: 데이터 메타 파일 내 라벨 파일 경로와 실제 라벨 파일 경로 일치 여부
        - 검사항목 007: 데이터 메타 파일의 이미지 내 객체 개수값(bbox_count)과 실제 라벨 파일 내 바운딩박스 개수의 일치 여부
        - 검사항목 008: 라벨 파일 내 바운딩박스 위치의 유효성 여부
        - 검사항목 009: 라벨 파일 내 바운딩박스 객체 범주값의 유효성

        parameter
        ----------
        filename(str): 파일 확장자를 제외한 이미지 파일명

        return
        ----------
        result(list): 검사항목 004, 006 ~ 011에 대한 검사 결과

        Usage
        ----------
        filename = 'Image_0001'
        result = check_img_exist_meta_table(filename = filename)
        """

        result_001 = self.ResultDict["QCinfo"].loc[self.ResultDict["QCinfo"]["파일명"] == filename, ["검사_001"]].values[0]

        # 검사항목 004, 006, 007
        if bool(result_001):

            meta_info = self.file_meta_table.loc[self.file_meta_table["파일명"] == filename, ["데이터셋구분", "라벨파일경로"]].values[0]
            meta_dataset_gubun, meta_label_path = meta_info

            # 검사항목 004: 메타 파일 내 데이터셋 구분값과 실제 이미지 파일 분류 현황 일치 여부
            current_img_path = self.PATH["ROOT"] / Path(self.ResultDict["QCinfo"].loc[self.ResultDict["QCinfo"]["파일명"] == filename, "이미지파일경로"].values[0])
            current_dataset_gubun = unicodedata.normalize("NFC", list(current_img_path.relative_to(self.PATH["IMAGE"]).parts)[0])
            result_004 = meta_dataset_gubun == current_dataset_gubun

            # 검사항목 005: 이미지 파일에 대한 라벨 파일 존재 여부
            try:
                current_label_path = self.PATH["ROOT"] / Path(self.ResultDict["QCinfo"].loc[self.ResultDict["QCinfo"]["파일명"] == filename, "라벨파일경로"].values[0])
                result_005 = current_label_path.exists()
            except:
                current_label_path = None
                result_005 = False

            try:
                # 검사항목 006: 데이터 메타 파일 내 라벨 파일 경로와 실제 라벨 파일 경로 일치 여부
                meta_label_path = self.PATH["ROOT"] / Path(meta_label_path)
            except:
                meta_label_path = None

            result_006 = meta_label_path == current_label_path

            try:
                # 라벨 파일 로드
                with open(current_label_path, "r") as f:
                    label_info = list(map(lambda x: x.strip().split(), f.readlines()))

                # 검사항목 007: 데이터 메타 파일의 이미지 내 객체 개수값(bbox_count)과 실제 라벨 파일 내 바운딩박스 개수의 일치 여부
                meta_bbox_count = self.file_meta_table.loc[self.file_meta_table["파일명"] == filename, "객체개수"].values[0]
                current_bbox_count = len(label_info)

                result_007 = meta_bbox_count == current_bbox_count

                # 검사항목 008: 라벨 파일 내 바운딩박스 객체 범주값의 유효성
                # 검사항목 009: 라벨 파일 내 바운딩박스 위치의 유효성 여부
                result_008, result_009 = True, True

                yaml_label_list = list(map(int, self.yaml_info["names"].keys()))

                for label_info_ in label_info:
                    label, x_center, y_center, width, height = label_info_

                    if int(label) not in yaml_label_list:
                        result_008 = False

                    if any([float(x_center) < 0, float(x_center) > 1, float(y_center) < 0, float(y_center) > 1, float(width) < 0, float(width) > 1, float(height) < 0, float(height) > 1, float(x_center) + (float(width) / 2) > 1, float(y_center) + (float(height) / 2) > 1]):
                        result_009 = False
            except:
                result_007, result_008, result_009 = False, False, False

        else:
            result_004, result_005, result_006, result_007, result_008, result_009 = False, False, False, False, False, False, False, False

        result = [int(result_004), int(result_005), int(result_006), int(result_007), int(result_008), int(result_009)]

        return result

    def QualityCheck(self):

        self.get_file_meta_table()

        self.logger.info("=" * 50)
        self.logger.info(f"{self.colorSetting['green']}데이터 QC 수행 시작{self.colorSetting['reset']}")

        self.img_filepath_list = [path for path in self.PATH["IMAGE"].rglob("*") if all([path.suffix.lower() in self.ext_fmt, not path.is_dir()])]  # 이미지 파일 경로 리스트 획득
        self.img_filename_list = [unicodedata.normalize("NFC", path.stem) for path in self.img_filepath_list]  # 이미지 파일명 리스트 획득

        max_digit_len = len(str(len(self.img_filepath_list)))  # No 컬럼값의 최대 자리수 파악
        QCinfo_init_df = pd.DataFrame(sorted([(unicodedata.normalize("NFC", img_path.stem), unicodedata.normalize("NFC", str(img_path.relative_to(self.PATH["ROOT"])))) if not isinstance(img_path.stem, int) else (unicodedata.normalize("NFC", str(img_path.stem)), unicodedata.normalize("NFC", str(img_path.relative_to(self.PATH["ROOT"])))) for img_path in self.img_filepath_list], key=lambda x: x[0]), columns=["파일명", "이미지파일경로"])

        if self.Mode == "객체탐지":
            self.label_filepath_list = [path for path in self.PATH["LABEL"].rglob("*") if all([path.suffix.lower() in [".txt"], not path.is_dir()])]  # [객체탐지] 어노테이션 파일 경로 리스트 획득
            self.label_filename_list = [unicodedata.normalize("NFC", path.stem) if not isinstance(path.stem, int) else unicodedata.normalize("NFC", str(path.stem)) for path in self.label_filepath_list]  # 이미지 파일명 리스트 획득
            self.QC_label_info = pd.DataFrame([(unicodedata.normalize("NFC", label_path.stem), unicodedata.normalize("NFC", str(label_path.relative_to(self.PATH["ROOT"])))) if not isinstance(label_path.stem, int) else (unicodedata.normalize("NFC", str(label_path.stem)), unicodedata.normalize("NFC", str(label_path.relative_to(self.PATH["ROOT"])))) for label_path in self.label_filepath_list], columns=["파일명", "라벨파일경로"])
            QCinfo_init_df = pd.merge(left=QCinfo_init_df, right=self.QC_label_info, on=["파일명"], how="left")

        self.ResultDict["QCinfo"] = pd.merge(left=QCinfo_init_df, right=self.file_meta_table[["파일명"]], on=["파일명"], how="left")
        self.ResultDict["QCinfo"] = self.ResultDict["QCinfo"].reset_index().rename(columns={"index": "No"})
        self.ResultDict["QCinfo"]["No"] = self.ResultDict["QCinfo"]["No"].apply(lambda x: str(x + 1).zfill(max_digit_len))

        # 공통 검사항목 검사 진행 (검사항목 001 ~ 003)
        # 검사항목 001: 데이터 메타 파일 존재 여부
        # 검사항목 002: 메타 파일 내 이미지 경로와 실제 이미지 경로 일치 여부
        # 검사항목 003: 메타 파일 내 이미지 해상도와 실제 이미지 해상도 일치 여부
        temp = pd.DataFrame([[filename] + self.check_common(filename) for filename in self.ResultDict["QCinfo"]["파일명"].values], columns=["파일명", "검사_001", "검사_002", "검사_003"])
        self.ResultDict["QCinfo"] = pd.merge(left=self.ResultDict["QCinfo"], right=temp, on=["파일명"], how="left")

        if self.Mode in ["분류"]:
            # 분류 검사항목 검사 진행 (검사항목 004 ~ 006)
            # 검사항목 004: 메타 파일 내 데이터셋 구분값과 실제 이미지 파일 분류 현황 일치 여부
            # 검사항목 005: 메타 파일 내 대분류값과 실제 이미지 파일 분류 현황 일치 여부
            # 검사항목 006: 메타 파일 내 소분류값과 실제 이미지 파일 분류 현황 일치 여부
            temp = pd.DataFrame([[filename] + self.check_classification(filename) for filename in self.ResultDict["QCinfo"]["파일명"].values], columns=["파일명", "검사_004", "검사_005"])
            self.ResultDict["QCinfo"] = pd.merge(left=self.ResultDict["QCinfo"], right=temp, on=["파일명"], how="left")
        elif self.Mode in ["객체탐지"]:
            # 객체탐지 검사항목 검사 진행 (검사항목 004 ~ 011)
            # 검사항목 004: 메타 파일 내 데이터셋 구분값과 실제 이미지 파일 분류 현황 일치 여부
            # 검사항목 005: 객체 범주 메타 파일(categories.json)
            # 검사항목 006: 메타 파일 내 어노테이션 파일 경로와 현재 어노테이션 파일 위치 일치 여부
            # 검사항목 007: 어노테이션 파일 내 이미지 파일 경로와 현재 이미지 파일 위치 일치 여부
            # 검사항목 008: 어노테이션 파일 내 이미지 해상도와 현재 이미지 파일 해상도 일치 여부
            # 검사항목 009: 어노테이션 파일 내 객체 개수값(bbox_count)과 바운딩박스 개수(len(bbox))의 일치 여부
            # 검사항목 010: 어노테이션 파일 내 바운딩박스 위치의 유효성 여부
            # 검사항목 011: 어노테이션 파일 내 바운딩박스 객체 범주값의 유효성
            temp = pd.DataFrame([[filename] + self.check_object_detection(filename) for filename in self.ResultDict["QCinfo"]["파일명"].values], columns=["파일명"] + [f"검사_{str(i).zfill(3)}" for i in range(4, 10)])
            self.ResultDict["QCinfo"] = pd.merge(left=self.ResultDict["QCinfo"], right=temp, on=["파일명"], how="left")

        self.logger.info(f"{self.colorSetting['green']}데이터 QC 수행 완료{self.colorSetting['reset']}")

        # self.ResultDict["QCinfo"].to_csv(self.PATH["OUTPUT"] / f"QualityCheck_Report_{self.Mode}.csv", index=False, encoding="utf-8-sig")

    def set_border(self, ws, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def Report(self):

        self.logger.info("=" * 50)
        self.logger.info(f"{self.colorSetting['green']}데이터 QC 결과 저장 작업 시작{self.colorSetting['reset']}")
        self.OutputPath = self.PATH["OUTPUT"] / f"QC결과서_{self.Mode}_{self.CreatedTime}.xlsx"

        # 결과서 정보 지정
        self.ResultDict["Report"]["제목"] = "데이터 품질 결과 보고서"
        self.ResultDict["Report"]["데이터셋명"] = self.DataName
        self.ResultDict["Report"]["활용구분"] = self.Mode
        self.ResultDict["Report"]["데이터수_이미지"] = len(self.img_filepath_list)
        if self.Mode == "일반":
            self.ResultDict["Report"]["검사항목"] = ["데이터 메타 파일 내 이미지 파일 정보 존재 여부", "데이터 메타 파일 내 이미지 경로와 실제 이미지 경로 일치 여부", "데이터 메타 파일 내 이미지 해상도와 실제 이미지 해상도 일치 여부"]
        elif self.Mode == "분류":
            self.ResultDict["Report"]["검사항목"] = ["데이터 메타 파일 내 이미지 파일 정보 존재 여부", "데이터 메타 파일 내 이미지 경로와 실제 이미지 경로 일치 여부", "데이터 메타 파일 내 이미지 해상도와 실제 이미지 해상도 일치 여부", "데이터 메타 파일 내 데이터셋 구분값과 실제 이미지 파일 분류 현황 일치 여부", "데이터 메타 파일 내 분류값과 실제 이미지 파일 분류 현황 일치 여부"]
        elif self.Mode == "객체탐지":
            self.ResultDict["Report"]["데이터수_어노테이션"] = len(self.label_filepath_list)
            self.ResultDict["Report"]["검사항목"] = ["데이터 메타 파일 내 이미지 파일 정보 존재 여부", "데이터 메타 파일 내 이미지 경로와 실제 이미지 경로 일치 여부", "데이터 메타 파일 내 이미지 해상도와 실제 이미지 해상도 일치 여부", "데이터 메타 파일 내 데이터셋 구분값과 실제 이미지 파일 분류 현황 일치 여부", "이미지 파일에 대한 라벨 파일 존재 여부", "데이터 메타 파일 내 라벨 파일 경로와 실제 라벨 파일 경로 일치 여부", "데이터 메타 파일의 이미지 내 객체 개수값(bbox_count)과 실제 라벨 파일 내 바운딩박스 개수의 일치 여부", "라벨 파일 내 바운딩박스 객체 범주값의 유효성", "라벨 파일 내 바운딩박스 위치값의 유효성 여부"]

        # 검사항목 별 검사 결과 종합
        check_result = [(idx, self.ResultDict["Report"]["검사항목"][idx - 1], int(self.ResultDict["QCinfo"].shape[0] == self.ResultDict["QCinfo"][f"검사_{str(idx).zfill(3)}"].sum()), f'{int(100 * self.ResultDict["QCinfo"][f"검사_{str(idx).zfill(3)}"].sum() / self.ResultDict["QCinfo"].shape[0])}% ({self.ResultDict["QCinfo"].shape[0]:,}개 중 {self.ResultDict["QCinfo"][f"검사_{str(idx).zfill(3)}"].sum():,}개) 일치') for idx in range(1, len(self.ResultDict["Report"]["검사항목"]) + 1)]
        check_result = pd.DataFrame(check_result, columns=["No", "검사 항목", "검사 결과", "비고"])
        check_result["검사 결과"] = check_result["검사 결과"].transform(lambda x: "O" if x == 1 else "X")

        check_cols = [f"검사_{str(i).zfill(3)}" for i in range(1, len(self.ResultDict["Report"]["검사항목"]) + 1)]
        for check_col in check_cols:
            self.ResultDict["QCinfo"][check_col] = self.ResultDict["QCinfo"][check_col].transform(lambda x: "O" if x == 1 else "X")

        with pd.ExcelWriter(self.OutputPath, mode="w", engine="openpyxl") as writer:

            # Step 6-1-a 테이블 리스트 시트
            check_result.to_excel(
                writer,
                index=False,
                header=True,
                sheet_name=f"결과서_{self.Mode}",
                startcol=0,
                startrow=5,
            )

            self.ResultDict["QCinfo"].to_excel(
                writer,
                index=False,
                header=True,
                sheet_name=f"파일 검사 결과",
                startcol=0,
                startrow=0,
            )

        wb = load_workbook(self.OutputPath)
        ws = wb[f"결과서_{self.Mode}"]

        # 결과서 정보 깂 입력
        end_row = 7 + check_result.shape[0]

        ws.cell(row=1, column=1, value=self.ResultDict["Report"]["제목"])
        ws.cell(row=2, column=1, value="데이터셋명")
        ws.cell(row=2, column=2, value=self.ResultDict["Report"]["데이터셋명"])
        ws.cell(row=3, column=1, value="활용 구분")
        ws.cell(row=3, column=2, value=self.ResultDict["Report"]["활용구분"])
        ws.cell(row=4, column=1, value="데이터 수")
        if self.Mode == "객체탐지":
            ws.cell(row=4, column=2, value=f'이미지파일: {self.ResultDict["Report"]["데이터수_이미지"]:,}개 / 어노테이션파일: {self.ResultDict["Report"]["데이터수_어노테이션"]:,}개')
        else:
            ws.cell(row=4, column=2, value=f'이미지파일: {self.ResultDict["Report"]["데이터수_이미지"]:,}')
        ws.cell(row=5, column=1, value="검사 결과")
        ws.cell(row=end_row, column=1, value="* 파일별 검사 결과는 [파일 검사 결과] 탭을 참고 바랍니다.")

        # 셀 병합 작업
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=4)
        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=4)
        ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=4)
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=4)
        ws.merge_cells(start_row=end_row, start_column=1, end_row=end_row, end_column=4)

        # 셀 높이, 너비 지정
        ws.row_dimensions[1].height = 34
        ws.column_dimensions["A"].width = 11
        ws.column_dimensions["B"].width = 65
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 30

        # 셀별 서식 지정
        thin = Side(style="thin")
        thick = Side(style="medium")

        ws.cell(row=1, column=1).font = Font(bold=True, size=16, underline="single")
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
        for col_i in range(1, 5):
            ws.cell(row=1, column=col_i).border = Border(top=thick, bottom=thick, right=thick, left=thick)

        for row_i in range(2, 5):
            ws.cell(row=row_i, column=1).font = Font(bold=True)
            ws.cell(row=row_i, column=1).fill = PatternFill(fgColor="ededed", fill_type="solid")
            ws.cell(row=row_i, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_i, column=1).border = Border(bottom=thin, right=thin, left=thick)

            ws.cell(row=row_i, column=2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_i, column=2).border = Border(bottom=thin, right=thin, left=thin)
            ws.cell(row=row_i, column=3).border = Border(bottom=thin, right=thin, left=thin)
            ws.cell(row=row_i, column=4).border = Border(bottom=thin, right=thick, left=thin)

        ws.cell(row=5, column=1).font = Font(bold=True)
        ws.cell(row=5, column=1).fill = PatternFill(fgColor="d9d9d9", fill_type="solid")
        ws.cell(row=5, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=5, column=1).border = Border(bottom=thin, right=thin, left=thick)
        ws.cell(row=5, column=2).border = Border(bottom=thin, right=thin, left=thin)
        ws.cell(row=5, column=3).border = Border(bottom=thin, right=thin, left=thin)
        ws.cell(row=5, column=4).border = Border(bottom=thin, right=thick, left=thin)

        for col_i in range(1, 5):
            ws.cell(row=6, column=col_i).font = Font(bold=True)
            ws.cell(row=6, column=col_i).fill = PatternFill(fgColor="ededed", fill_type="solid")
            ws.cell(row=6, column=col_i).alignment = Alignment(horizontal="center", vertical="center")
            if col_i == 1:
                ws.cell(row=6, column=col_i).border = Border(bottom=thin, right=thin, left=thick)
            elif col_i == 4:
                ws.cell(row=6, column=col_i).border = Border(bottom=thin, right=thick, left=thin)
            else:
                ws.cell(row=6, column=col_i).border = Border(bottom=thin, right=thin, left=thin)

        for row_i in range(7, 7 + check_result.shape[0]):
            ws.cell(row=row_i, column=1).font = Font(bold=True)
            ws.cell(row=row_i, column=1).fill = PatternFill(fgColor="d9d9d9", fill_type="solid")
            ws.cell(row=row_i, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_i, column=1).border = Border(bottom=thin, right=thin, left=thick)

            ws.cell(row=row_i, column=2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_i, column=2).border = Border(bottom=thin, right=thin, left=thin)

            ws.cell(row=row_i, column=3).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_i, column=3).border = Border(bottom=thin, right=thin, left=thin)

            ws.cell(row=row_i, column=4).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_i, column=4).border = Border(bottom=thin, right=thick, left=thin)

        ws.cell(row=end_row, column=1).font = Font(bold=True)
        ws.cell(row=end_row, column=1).fill = PatternFill(fgColor="ededed", fill_type="solid")
        ws.cell(row=end_row, column=1).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=end_row, column=1).border = Border(bottom=thin, right=thin, left=thick)
        ws.cell(row=end_row, column=2).border = Border(bottom=thin, right=thin, left=thin)
        ws.cell(row=end_row, column=3).border = Border(bottom=thin, right=thin, left=thin)
        ws.cell(row=end_row, column=4).border = Border(bottom=thin, right=thick, left=thin)

        cond_cells = f"C7:C{end_row-1}"
        success_font = Font(size=12, color="0000ff")
        success_fill = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid")

        fail_font = Font(size=12, color="ff0000")
        fail_fill = PatternFill(start_color="fbc7ce", end_color="fbc7ce", fill_type="solid")

        ws.conditional_formatting.add(cond_cells, formatting.rule.FormulaRule(formula=['NOT(ISERROR(SEARCH("O",C7)))'], stopIfTrue=True, fill=success_fill, font=success_font))
        ws.conditional_formatting.add(cond_cells, formatting.rule.FormulaRule(formula=['NOT(ISERROR(SEARCH("X",C7)))'], stopIfTrue=True, fill=fail_fill, font=fail_font))

        # 파일 검사 결과 시트 서식 지정
        ws = wb["파일 검사 결과"]

        start_col = 0
        val_check = False
        while not val_check:
            if ws.cell(row=1, column=start_col + 1).value == "검사_001":
                val_check = True
            else:
                start_col += 1
        end_col = ws.max_column
        start_row = 2
        end_row = ws.max_row

        self.set_border(ws, f"A1:{chr(65 + end_col - 1)}{end_row}")

        for col_i in range(1, end_col + 1):
            ws.cell(row=1, column=col_i).fill = PatternFill(fgColor="ededed", fill_type="solid")
            ws.cell(row=1, column=col_i).alignment = Alignment(horizontal="center", vertical="center")

        cond_cells = f"{chr(65 + start_col)}{start_row}:{chr(65 + end_col - 1)}{end_row}"
        success_font = Font(size=12, color="0000ff")
        success_fill = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid")

        fail_font = Font(size=12, color="ff0000")
        fail_fill = PatternFill(start_color="fbc7ce", end_color="fbc7ce", fill_type="solid")

        ws.conditional_formatting.add(cond_cells, formatting.rule.FormulaRule(formula=[f'NOT(ISERROR(SEARCH("O",{chr(65 + start_col)}{start_row})))'], stopIfTrue=True, fill=success_fill, font=success_font))
        ws.conditional_formatting.add(cond_cells, formatting.rule.FormulaRule(formula=[f'NOT(ISERROR(SEARCH("X",{chr(65 + start_col)}{start_row})))'], stopIfTrue=True, fill=fail_fill, font=fail_font))

        wb.save(self.OutputPath)

        self.logger.info("=" * 50)
        self.logger.info(f"{self.colorSetting['green']}데이터 QC 결과 저장 작업 완료{self.colorSetting['reset']}")

    def close(self):
        self.timer.stop()
