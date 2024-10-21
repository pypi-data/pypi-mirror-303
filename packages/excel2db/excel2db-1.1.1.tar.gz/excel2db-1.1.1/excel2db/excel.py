# -*- coding:utf8 -*-
"""
##执行excel级别转换
"""
import os, xlrd, csv
from pandas import read_excel, read_csv, errors, DataFrame
from openpyxl import load_workbook
from . import cheakConf, sheet, insert2sqlite

class excel:
    def __init__(self, value, conf):

        """
        excel级别操作
        :param value: 变量文件
        :param conf: excel级别配置（清洗前）
        :param excelUrl: excel文件路径
        """
        self.value = value
        cheakConf.excelConf(self.value, conf) ##获取excel级别配置(清洗后)

    def excel(self, excelUrl):
        ##获取excel文件名
        (path, fileName) = os.path.split(excelUrl)
        (name, suffix) = os.path.splitext(fileName)
        self.value.excelUrl = excelUrl
        self.value.excelSuffix = suffix
        self.value.excelFileName = name

        ##获取类型小写
        suffixLower = suffix.lower()

        ##打开excel文件
        if suffixLower == '.xls':
            df = self.__readXls()
        elif suffixLower == '.xlsx':
            df = self.__readXlsx()
        elif suffixLower == '.csv':
            df = self.__readCsv()
        else:
            raise Exception("文件类型错误")

        self.value.sheetList = list(df.keys())
        self.value.excelData = df

        if self.value.excelConf["isSaveDatabase"] == True:
            self.value.dbClass = insert2sqlite.insert2sqlite(self.value)

        sheetConf = cheakConf.combinSheetConf(self.value) ##sheet级别配置文件合并

        for conf in sheetConf:
            sheetManager = sheet.sheet(self.value, conf)
            sheetManager.sheet()

    def __readXls(self):
        if self.value.excelConf["notReadHideSheet"] == True:
            sheetName = self.__read_visible_sheets_xls(self.value.excelUrl)
        else:
            sheetName = None
        df = read_excel(
            io=self.value.excelUrl,
            sheet_name=sheetName,
            header=None
        )

        return df

    def __readXlsx(self):
        if self.value.excelConf["notReadHideSheet"] == True:
            sheetName = self.__read_visible_sheets_xlsx(self.value.excelUrl)
        else:
            sheetName = None
        df = read_excel(
            io=self.value.excelUrl,
            sheet_name=sheetName,
            header=None
        )

        return df

    def __readCsv(self):
        ## csv文件中没有sheet概念，因此无需进行多sheet操作
        try:
            df = read_csv(
                filepath_or_buffer=self.value.excelUrl,
                header=None
            )
        except errors.ParserError as e: ##若识别失败，尝试手动识别
            data = []
            with open(self.value.excelUrl, mode='r', newline='', encoding='utf-8') as file:
                csv_reader = csv.reader(file)
                for row in csv_reader:
                    data.append(row)
            df = DataFrame(data)

        df = {self.value.excelFileName: df}

        return df

    def __read_visible_sheets_xls(self, filename):
        # 打开 Excel 文件
        book = xlrd.open_workbook(filename)

        # 获取所有工作表的名称
        all_sheet_names = book.sheet_names()

        # 过滤出未隐藏的工作表名称
        visible_sheet_names = [name for name in all_sheet_names if book.sheet_by_name(name).visibility == 0]

        return visible_sheet_names

    def __read_visible_sheets_xlsx(self, filename):
        # 使用 openpyxl 加载 Excel 文件
        wb = load_workbook(filename, read_only=True)

        # 获取所有工作表的名称
        all_sheet_names = wb.sheetnames

        # 过滤出未隐藏的工作表名称
        visible_sheet_names = [name for name in all_sheet_names if not wb[name].sheet_state == 'hidden']

        return visible_sheet_names