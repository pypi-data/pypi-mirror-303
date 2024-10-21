import os
from typing import Set, Tuple, Optional, Union, List, Dict

import openpyxl
from openpyxl.workbook import workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

from dubna_parser.downloader import download_sheets
from dubna_parser.help_functions import join_dicts
from dubna_parser.models import GroupModel, PairModel, AlternatingPairModel, SchedulePairModel, GroupPairsScheduleModel, \
    ScheduleModel


class ScheduleParser:
    def __init__(self):
        self.save_folder = 'downloads'
        self.file_with_url = 'links'
        self.specializations_with_groups: Dict[str, List[GroupModel]] = dict()
        self.pairs_of_groups: List[GroupPairsScheduleModel] = list()
        self.specializations: Set[str] = set()
        self.groups: Set[GroupModel] = set()
        self.classrooms: Set[str] = set()
        self.subjects: Set[str] = set()
        self.teachers: Set[str] = set()
        self.all_specializations: Set[str] = set()
        # немного не знаю куда это пихать
        self.degrees = ['проф.', 'доцент', 'доц.', ' ст.пр.', 'пр.', 'ст.преподаватель', 'ст.преп.', 'профессор']

    # методы для моделей
    def get_model_pair(self, classroom: str,
                       subject: str,
                       teacher: str) -> PairModel:
        self.classrooms.add(classroom)
        self.subjects.add(subject)
        self.teachers.add(teacher)
        return PairModel(classroom=classroom,
                         subject=subject,
                         teacher=teacher)

    def serialize_group(self, group, specialization: str) -> GroupModel:
        group_number = 0
        if group.isdigit():
            group_number = int(group)

        group_model = GroupModel(group=group_number,
                                 specialization=specialization)
        self.groups.add(group_model)
        return group_model

    @property
    def save_folder(self):
        return self._save_folder

    @save_folder.setter
    def save_folder(self, value: str):
        self._save_folder = value

    @property
    def file_with_url(self):
        return self._file_with_url

    @file_with_url.setter
    def file_with_url(self, value: str):
        self._file_with_url = value

    def set_params_for_download(self, file_with_url: Optional[str], save_folder: Optional[str]):
        if file_with_url:
            self.file_with_url = file_with_url
        if save_folder:
            self.save_folder = save_folder
        if (not self.file_with_url) or (not self.save_folder):
            raise Exception("Please provide either file_with_url and file_with_url")

    def download(self, file_with_url: Optional[str], save_folder: Optional[str]):
        self.set_params_for_download(file_with_url, save_folder)
        download_sheets(file_with_url, save_folder)

    def get_table_names(self, wb: workbook.Workbook) -> List[str]:
        return wb.sheetnames

    def get_not_empty_columns(self, ws: Worksheet, index_start=3) -> (int, int):
        current_row = 4
        not_empty_column_index = index_start
        for index in range(index_start, ws.max_column + 1):
            if ws.cell(row=current_row, column=not_empty_column_index).value is None:
                break
            not_empty_column_index += 1
        return index_start, not_empty_column_index

    def get_specializations_from_row(self, ws: Worksheet, index_start=3,
                                     index_end=3, index_row=4) -> Dict[str, List[GroupModel]]:
        groups: Dict[str, List[GroupModel]] = dict()
        for col in ws.iter_cols(min_col=index_start, max_col=index_end,
                                min_row=index_row, max_row=index_row,
                                values_only=True):
            for cell in col:
                if cell:
                    group_with_specialization = cell.strip().split(' ', 1)
                    if len(group_with_specialization) == 1:
                        number_group, special = cell.strip().split('(', 1)
                        group_with_specialization = number_group, special
                    specialization = group_with_specialization[1].strip("()")
                    group = group_with_specialization[0]

                    group = self.serialize_group(group, specialization)

                    if specialization in groups:
                        groups[specialization].append(group)
                    else:
                        new_group = list()
                        new_group.append(group)
                        groups[specialization] = new_group
        return groups

    def get_indexes_of_weeks(self, ws: Worksheet, column_index=1):
        day_indices = dict()
        for crange in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = crange.bounds
            if min_col == column_index and max_col == column_index:
                day_indices[ws.cell(row=min_row, column=min_col).value] = (min_row, max_row)
        return day_indices

    def get_pair_from_row(self, row: str) -> PairModel:
        row = row.replace('/', '').strip()
        for degree in self.degrees:
            row = row.replace(degree, '').strip()
        classroom = ''
        teacher = ''
        for letter in row:
            if letter != ' ':
                classroom += letter
            else:
                break
        space_count = 0
        rev_row = row[::-1]
        for letter in rev_row:
            if space_count >= 2:
                break
            if letter != ' ':
                teacher += letter
            else:
                teacher += ' '
                space_count += 1
        teacher = teacher[::-1].strip()

        subject = row.replace(teacher, '').replace(classroom, '').strip()

        return self.get_model_pair(classroom, subject, teacher)

    def get_rows(self, ws: Worksheet, column_index,
                 first_row_index, second_row_index) -> Tuple[str, str]:
        row1, row2 = '', ''
        for crange in ws.merged_cells.ranges:
            if row1 and row2:
                break
            min_col, min_row, max_col, max_row = crange.bounds
            if min_col <= column_index <= max_col and min_row == first_row_index:
                row1 = ws.cell(row=min_row, column=min_col).value
            if min_col <= column_index <= max_col and min_row == second_row_index:
                row2 = ws.cell(row=min_row, column=min_col).value
        else:
            if row1:
                row2 = ws.cell(row=second_row_index, column=column_index).value
            elif row2:
                row1 = ws.cell(row=first_row_index, column=column_index).value
            else:
                row1 = ws.cell(row=first_row_index, column=column_index).value
                row2 = ws.cell(row=second_row_index, column=column_index).value
        return row1, row2

    def get_single_row(self, ws: Worksheet, column_index, row_index) -> str:
        row = ''
        for crange in ws.merged_cells.ranges:
            if row:
                break
            min_col, min_row, max_col, max_row = crange.bounds
            if min_col <= column_index <= max_col and min_row == row_index:
                row = ws.cell(row=min_row, column=min_col).value
        else:
            row = ws.cell(row=row_index, column=column_index).value
        return row

    def get_pair(self, row1, row2) -> Optional[Union[PairModel, AlternatingPairModel]]:
        pair = None
        if row1 == 'с/к Олимп':
            classroom = row1
            subject = row2
            teacher = 'undefined'
            pair = PairModel(classroom=classroom, subject=subject, teacher=teacher)
        elif row1 and row2:
            if row1.count('/') >= 1 and row2.count('/') >= 1:
                pair1 = self.get_pair_from_row(row1)
                pair2 = self.get_pair_from_row(row2)
                pair = AlternatingPairModel(odd_week=pair1, even_week=pair2)
                # пара мигалка
            else:
                row = row1 + '  ' + row2
                pair = self.get_pair_from_row(row)
        elif row1:
            pair1 = self.get_pair_from_row(row1)
            if row1.count('/') == 0:
                pair = pair1
            if row1.count('/') > 0:
                pair = AlternatingPairModel(odd_week=pair1, even_week=None)
        elif row2:
            pair2 = self.get_pair_from_row(row2)
            if row2.count('/') == 0:
                pair = pair2
            if row2.count('/') > 0:
                pair = AlternatingPairModel(odd_week=None, even_week=pair)
        else:
            pair = None
        return pair

    def get_single_pair(self, row) -> Optional[Union[PairModel, AlternatingPairModel]]:
        pair = None
        if row:
            row.strip()
            if row.count('/') == 0:
                pair = self.get_pair_from_row(row)
            else:
                if row[0] == '/':
                    pair2 = self.get_pair_from_row(row)
                    pair = AlternatingPairModel(odd_week=None, even_week=pair2)
                elif row[-1] == '/':
                    pair1 = self.get_pair_from_row(row)
                    pair = AlternatingPairModel(odd_week=pair1, even_week=None)
                else:
                    print("тут уже мне неизвестно как парсить...")
        return pair

    def get_pairs_for_group(self, ws: Worksheet,
                            indices_of_week, column_index: int) \
            -> Dict[str, List[SchedulePairModel]]:
        group_pairs: Dict[str, List[SchedulePairModel]] = dict()
        expected_day = 'вторник'
        for week_day, day_index in indices_of_week.items():
            start_day_index, end_day_index = day_index
            pairs_for_day: List[SchedulePairModel] = list()
            for i in range(start_day_index, end_day_index, 2):
                first_row_index, second_row_index = i, i + 1
                row1, row2 = self.get_rows(ws, column_index, first_row_index, second_row_index)
                pair_number = (i - start_day_index) // 2 + 1
                pair = self.get_pair(row1, row2)
                schedule_pair = SchedulePairModel(pair_number=pair_number, pair=pair)
                pairs_for_day.append(schedule_pair)

            if (end_day_index - start_day_index) % 2 == 1:
                # осталась последняя пара с одной строкой
                row = self.get_single_row(ws, column_index, end_day_index)
                pair_number = (end_day_index - start_day_index) // 2 + 1
                pair = self.get_single_pair(row)
                pair = SchedulePairModel(pair_number=pair_number, pair=pair)
                pairs_for_day.append(pair)
            if week_day:
                group_pairs[week_day] = pairs_for_day
            else:
                group_pairs[expected_day] = pairs_for_day
        return group_pairs

    def set_default_data(self):
        self.specializations_with_groups: Dict[str, List[GroupModel]] = dict()
        self.pairs_of_groups: List[GroupPairsScheduleModel] = list()
        self.specializations: Set[str] = set()
        self.groups: Set[GroupModel] = set()
        self.classrooms: Set[str] = set()
        self.subjects: Set[str] = set()
        self.teachers: Set[str] = set()
        self.all_specializations: Set[str] = set()

    def parse(self, save_folder: Optional[str]) -> ScheduleModel:
        """перед парсингом не открывать таблицы! Ничего не трогать!"""
        self.set_default_data()
        all_files = os.listdir(save_folder)
        for filename in all_files:
            current_file = os.path.join(save_folder, filename)
            excel_file = openpyxl.load_workbook(current_file)
            names = self.get_table_names(excel_file)
            specializations_with_groups_from_file = dict()
            for name in names:
                current_table = excel_file[name]
                indices_of_week = self.get_indexes_of_weeks(current_table)
                group_indexes = self.get_not_empty_columns(current_table)
                start_index, end_index = group_indexes
                specializations = self.get_specializations_from_row(current_table, start_index, end_index)

                all_values = [item for sublist in specializations.values() for item in sublist]
                for group_index, group in zip(range(start_index, end_index), all_values):
                    pairs_for_group = self.get_pairs_for_group(current_table, indices_of_week, group_index)
                    try:
                        pair_of_group = GroupPairsScheduleModel(group=group, schedule_pairs=pairs_for_group)
                        self.pairs_of_groups.append(pair_of_group)
                    except ValidationError as e:
                        print(e)

                specializations_with_groups_from_file = join_dicts(specializations_with_groups_from_file,
                                                                   specializations)
            self.specializations_with_groups = join_dicts(self.specializations_with_groups,
                                                          specializations_with_groups_from_file)
        schedule = ScheduleModel(
            specializations_with_groups=self.specializations_with_groups,
            specializations=self.specializations,
            groups=self.groups,
            classrooms=self.classrooms,
            teachers=self.teachers,
            subjects=self.subjects,
            schedule_pairs=self.pairs_of_groups
        )
        return schedule
